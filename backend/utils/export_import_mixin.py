# -*- coding: utf-8 -*-
# +-------------------------------------------------------------------
# | dvlyadmin-mini
# +-------------------------------------------------------------------
# | Author: lybbn
# +-------------------------------------------------------------------
# | QQ: 1042594286
# +-------------------------------------------------------------------
# | EditDate: 2025-06-20
# +-------------------------------------------------------------------
# | 名称: 导入导出功能
# +-------------------------------------------------------------------
# | 版本: 1.0
# +-------------------------------------------------------------------

# ------------------------------
# 导出excel数据，支持导出图片(单张url网络地址),支持嵌套(1层)数据 如 ： {'a':'b','c':{'d':'e'}} 要获取d的值，field_data 直接{'c.d':"嵌套1层获取"}即可
# 网络图片不存在会直接写入图片url网络地址
# ------------------------------

import os
import time
import random
from io import BytesIO
import requests
from django.conf import settings
from utils.common import getfulldomian
from rest_framework.request import Request
from urllib.request import pathname2url
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from django.utils.encoding import escape_uri_path
from django.http import FileResponse, HttpResponse
from django.core.validators import URLValidator
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from utils.jsonResponse import SuccessResponse,DetailResponse,ErrorResponse
from django.db import transaction
from openpyxl.styles import Alignment, Font, PatternFill
import logging

logger = logging.getLogger(__name__)

class ImportExportMixin:
    """
    导入导出功能Mixin
    支持功能：
    - 导出Excel（支持图片、嵌套字段）
    - 导入Excel（支持嵌套字段映射）
    - 下载导入模板
    配置参数:
    - import_field_dict: 导入字段映射 {'excel列名': 'model字段名'}，支持嵌套如 'c.d'
    - export_field_dict: 导出字段标签 {'model字段名': 'excel列名'}，支持嵌套如 'c.d'
    - import_image_fields: 需要特别处理的图片字段列表 ['avatar', 'logo']
    """
    default_file_name = "ly"+time.strftime('%Y%m%d%H%M%S')+ str(random.randint(10, 99))+".xlsx"
    import_field_dict = {}
    export_field_dict = {}
    import_image_fields = []
    # 表格表头最大宽度，默认80个字符
    export_column_width = 80
    # 保存文件名
    file_name = default_file_name
    # 保存目录
    save_dir = os.path.join('systemexport', time.strftime('%Y-%m-%d', time.localtime(time.time())))
    # 保存位置
    save_path = os.path.join(settings.MEDIA_ROOT, save_dir, file_name)

    # 是否显示图片,否的话为地址
    show_image = True

    def is_image(self, urlstr):
        """是否为图片地址"""
        val = URLValidator(schemes=('http', 'https'))
        try:
            val(urlstr)
            white_list = ['.png','.jpg','.jpeg','.gif']
            a = urlparse(urlstr)
            file_path = a.path
            file_name = os.path.basename(file_path)
            _,file_suffix = os.path.splitext(file_name)
            if file_suffix.lower() in white_list:
                return True
            return False
        except:
            return False

    def is_number(self, num):
        """是否为数字"""
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """获取字符串最大长度"""
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            string = str(string)
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    def pixels_to_points(self, value, dpi=96):
        """96 dpi, 72i"""
        return value * 72 / dpi

    def inserImg(self, sheet, Imgpath, row_index, col_index):
        """插入图片"""
        try:
            response = requests.get(Imgpath)
            img = BytesIO(response.content)
            imgToInsert = Image(img)
            oriImgH, oriImgW = imgToInsert.height, imgToInsert.width
            resize_factor = 0.6
            w_h_ratio = oriImgW/oriImgH
            resize_H = 40
            resize_W = 40
            imgsize_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
            cell_height = int(self.pixels_to_points(resize_H+10, dpi=96))
            col_letter = get_column_letter(col_index)
            x_pad = 8
            y_pad = 5
            marker = AnchorMarker(col=col_index-1, colOff=pixels_to_EMU(x_pad), row=row_index-1, rowOff=pixels_to_EMU(y_pad))
            imgToInsert.anchor = OneCellAnchor(_from=marker, ext=imgsize_excel)
            sheet.add_image(imgToInsert)
            return sheet
        except Exception as e:
            logger.error(f"插入图片失败: {Imgpath}, 错误: {str(e)}")
            return sheet

    def cell_nomall_data(self, ws, index, h_index, df_len_max, val, left_center):
        """处理cell正常数据"""
        if isinstance(val, list) and not val:
            val = ""
        if isinstance(val, dict) or isinstance(val, list):
            val = str(val)
        ws.cell(row=index+2, column=h_index+1).value = val
        ws.cell(row=index+2, column=h_index+1).alignment = left_center
        result_column_width = self.get_string_len(val)
        if h_index != 0 and result_column_width > df_len_max[h_index]:
            df_len_max[h_index] = result_column_width
        return ws

    def cell_image_data(self, ws, index, h_index, df_len_max, val):
        """处理cell图像插入"""
        self.inserImg(ws, val, index+2, h_index+1)
        if h_index != 0 and 10 > df_len_max[h_index]:
            df_len_max[h_index] = 10
        return ws

    def write_cell_data(self, ws, index, h_index, df_len_max, val, left_center):
        """处理cell插入"""
        if self.show_image and self.is_image(val):
            try:
                self.inserImg(ws, val, index+2, h_index+1)
                if h_index != 0 and 10 > df_len_max[h_index]:
                    df_len_max[h_index] = 10
            except:
                self.cell_nomall_data(ws, index, h_index, df_len_max, val, left_center)
        else:
            self.cell_nomall_data(ws, index, h_index, df_len_max, val, left_center)
        return ws

    def get_export_queryset(self):
        """获取导出数据的queryset，可被子类重写"""
        return self.filter_queryset(self.get_queryset())

    def before_import_row(self, row_data, row_num):
        """
        导入前处理单行数据
        可被子类重写用于数据验证或转换
        """
        return row_data

    def after_import(self, success_count, error_count, error_messages):
        """
        导入完成后处理
        可被子类重写
        """
        pass

    def process_nested_field(self, field_name, value, data_dict, is_many=False, separator=','):
        """
        处理嵌套字段和多对多字段
        
        参数:
            field_name: 字段名，可以是普通字段或嵌套字段(如 'user.name')
            value: 字段值
            data_dict: 要填充的数据字典
            is_many: 是否多对多字段
            separator: 多值分隔符
        """
        if not field_name:
            return
        
        # 处理空值
        if value is None:
            value = ""
        value = str(value).strip()
        
        # 处理多对多字段
        if is_many:
            try:
                related_ids = []
                if value:
                    # 分割多值
                    values = [v.strip() for v in value.split(separator) if v.strip()]
                    
                    for val in values:
                        related_ids.append(val)
                
                # 如果是嵌套的多对多字段
                if '.' in field_name:
                    parts = field_name.split('.')
                    current = data_dict
                    for part in parts[:-1]:
                        if part not in current:
                            current[part] = {}
                        current = current[part]
                    current[parts[-1]] = related_ids
                else:
                    data_dict[field_name] = related_ids
                return
            except Exception as e:
                raise ValueError(f"处理多对多字段'{field_name}'错误: {str(e)}")
        
        # 处理普通嵌套字段
        if '.' in field_name:
            parts = field_name.split('.')
            current = data_dict
            for part in parts[:-1]:
                if part not in current:
                    current[part] = {}
                current = current[part]
            current[parts[-1]] = value
        else:
            data_dict[field_name] = value
    
    def get_export_filename(self,queryset):
        """
        取导出文件名
        """
        model = None
        if queryset is not None and hasattr(queryset, 'model'):
            model = queryset.model
        if model:
            return getattr(model, '_meta').verbose_name
        else:
            model = queryset.model._meta.verbose_name
        if not model:return self.file_name
        return model

    @action(detail=False, methods=['post'])
    def export_data(self, request, *args, **kwargs):
        """
        导出数据为Excel（支持图片和嵌套字段）,支持前端自定义字段和勾选数据导出
        @param: export_fields 优先使用前端传递的导出字段
        @param: selected_ids 只导出前端勾选出来的数据id列表
        请求体示例:
        {
            "selected_ids": [1, 2, 3],
            "export_fields": {
                "name": "产品名称",
                "code": "产品代码"
            }
        }
        """
        export_fields = request.data.get('export_fields', {})
        if export_fields: # 前端传递了导出字段，使用前端传递的导出字段
            self.export_field_dict = export_fields
        selected_ids = request.data.get('selected_ids', [])
        queryset = self.get_export_queryset()
        if selected_ids:
            queryset = queryset.filter(id__in=selected_ids)
        if not self.export_serializer_class:self.export_serializer_class = self.serializer_class
        assert self.export_serializer_class, "'%s' 请配置对应的导出序列化器。" % self.__class__.__name__
        serializer = self.export_serializer_class(queryset, many=True, request=request)
        data = serializer.data
        
        if not data:
            return ErrorResponse(msg='没有可导出的数据')
        
        try:
            wb = Workbook()
            ws = wb.active
            left_center = Alignment(horizontal='left', vertical='center')
            left_center_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 处理字段映射
            field_order = ["#"] + list(self.export_field_dict.keys())
            header_data = ["序号"] + list(self.export_field_dict.values())
            
            df_len_max = [self.get_string_len(i) for i in header_data]
            row = get_column_letter(len(self.export_field_dict) + 1)
            column = 1
            
            # 写入表头
            ws.append(header_data)
            
            # 写入数据
            for index, item in enumerate(data):
                row_data = [index + 1]  # 序号列
                for field in field_order[1:]:  # 跳过序号列
                    if '.' in field:  # 处理嵌套字段
                        parts = field.split('.')
                        value = item
                        try:
                            for part in parts:
                                value = value[part]
                            # 处理数组/列表类型值
                            if isinstance(value, (list, tuple)):
                                value = ",".join(str(v) for v in value)
                            row_data.append(str(value) if value is not None else "")
                        except (KeyError, TypeError):
                            row_data.append("")
                    else:
                        value = item.get(field, "")
                        # 处理数组/列表类型值
                        if isinstance(value, (list, tuple)):
                            value = ", ".join(str(v) for v in value)
                        row_data.append(str(value) if value is not None else "")
                
                # 写入一行数据
                ws.append(row_data)
                
                # 处理特殊字段（图片等）
                for h_index, h_item in enumerate(field_order):
                    if h_index == 0:
                        ws.cell(row=index+2, column=h_index+1).alignment = left_center
                    else:
                        field_name = h_item
                        cell_value = row_data[h_index]
                        
                        # 如果是图片字段且值是URL
                        if (field_name in self.import_image_fields or 
                            ('.' in field_name and field_name.split('.')[-1] in self.import_image_fields)):
                            if cell_value and self.is_image(cell_value):
                                try:
                                    self.inserImg(ws, cell_value, index+2, h_index+1)
                                    if 10 > df_len_max[h_index]:
                                        df_len_max[h_index] = 10
                                except:
                                    pass
                        else:
                            # 普通字段处理
                            self.cell_nomall_data(ws, index, h_index, df_len_max, cell_value, left_center)
                
                column += 1
            
            # 更新列宽
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            
            # 添加过滤头
            tab = Table(displayName="Table", ref=f"A1:{row}{column}")
            style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=True, showLastColumn=True, showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # 返回文件
            output_buffer = BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            response = HttpResponse(
                output_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            export_file_name = f'导出{self.get_export_filename(queryset)}数据.xlsx'
            response['Content-Disposition'] = f'attachment; filename={escape_uri_path(export_file_name)}'
            return response
            
        except Exception as e:
            logger.error(f"导出失败: {str(e)}")
            return ErrorResponse(msg=f'导出失败: {str(e)}')
    
    @action(detail=False, methods=['post'])
    @transaction.atomic
    def import_data(self, request, *args, **kwargs):
        """批量导入Excel数据（支持嵌套字段）"""
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        assert self.import_serializer_class, "'%s' 请配置对应的导入序列化器。" % self.__class__.__name__
        
        if 'file' not in request.FILES:
            return ErrorResponse(msg='请上传文件')
        
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return ErrorResponse(msg='仅支持Excel文件')
        
        try:
            wb = load_workbook(filename=BytesIO(file.read()))
            ws = wb.active
            
            # 获取表头映射
            headers = [cell.value for cell in ws[1]]
            
            # 验证表头
            missing_headers = set(self.import_field_dict.keys()) - set(headers)
            if missing_headers:
                return ErrorResponse(msg=f'缺少必要的列: {", ".join(missing_headers)}')
            
            # 准备批量数据
            valid_data = []
            error_count = 0
            error_messages = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                
                try:
                    row_data = {}
                    for idx, cell_value in enumerate(row):
                        if idx >= len(headers):
                            continue
                        
                        excel_col_name = headers[idx]
                        if excel_col_name in self.import_field_dict:
                            field_info = self.import_field_dict[excel_col_name]
                            model_field = field_info if isinstance(field_info, str) else field_info['field']

                            # 获取字段配置
                            is_many = False
                            separator = ","
                            if isinstance(field_info, dict):
                                is_many = field_info.get('many', False)
                                separator = field_info.get('separator', ',')
                            
                            # 处理图片字段
                            if hasattr(self, 'import_image_fields') and (
                                model_field in self.import_image_fields or 
                                ('.' in model_field and model_field.split('.')[-1] in self.import_image_fields)):
                                cell_value = cell_value if (cell_value and self.is_image(cell_value)) else None
                            
                            # 处理普通字段
                            cell_value = "" if cell_value is None else cell_value
                            self.process_nested_field(model_field, cell_value, row_data, is_many, separator)
                    
                    # 预处理
                    processed_data = self.before_import_row(row_data, row_idx)
                    valid_data.append(processed_data)
                    
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"第{row_idx}行错误: {str(e)}")
                    continue
            
            # 批量导入
            if valid_data:
                serializer = self.import_serializer_class(data=valid_data, many=True)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            
            # 后处理
            self.after_import(len(valid_data), error_count, error_messages)
            
            return DetailResponse(data={
                'success': True,
                'success_count': len(valid_data),
                'error_count': error_count,
                'error_messages': error_messages
            })
            
        except Exception as e:
            logger.error(f"导入失败: {str(e)}", exc_info=True)
            return ErrorResponse(msg=f'导入失败: {str(e)}')
    
    def generate_template(self):
        """
        生成简化版导入模板（仅包含字段标题和示例数据）
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        
        # 基础样式设置
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        left_center = Alignment(horizontal='left', vertical='center')
        
        # 准备数据
        headers = []
        example_values = []
        
        # 获取示例数据
        example_data = self.get_import_example_data() if hasattr(self, 'get_import_example_data') else {}
        
        for excel_col, field_info in self.import_field_dict.items():
            # 支持字典和字符串两种配置方式
            if isinstance(field_info, dict):
                model_field = field_info.get('field', field_info)
                example = field_info.get('example', example_data.get(model_field, ''))
            else:
                example = example_data.get(field_info, '')
            
            headers.append(excel_col)
            example_values.append(example)
        
        # 写入表头行
        ws.append(headers)
        
        # 写入示例数据行
        ws.append(example_values)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_center
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20
        
        # 冻结表头行
        ws.freeze_panes = "A2"
        
        # 返回文件流
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @action(detail=False, methods=['post'])
    def download_template(self, request, *args, **kwargs):
        """
        下载导入模板
        """
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        try:
            template_stream = self.generate_template()
            response = HttpResponse(
                template_stream.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            mname = self.get_export_filename(self.get_export_queryset())
            filename = f"import_template_{mname}.xlsx" if mname else f"import_template_{time.strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename={escape_uri_path(filename)}'
            return response
        except Exception as e:
            logger.error(f"生成模板失败: {str(e)}")
            return ErrorResponse(msg=f'生成模板失败: {str(e)}')
